import os
from os.path import isfile, join
from pathlib import Path
import sys
import re
from shutil import copyfile
from datetime import date
from openpyxl import load_workbook

PROGRAM_ANNOUNCEMENT_STRING = """
[ 날짜 입력기 (ver.공무일지) ]

"""
EXCEL_EXTENSIONS = ["xlsx", "xlsm", "xlsb", "xltx", "xls", "xlt", "xml", "xlam", "xlw", "xlr"]

current_path = ""
excel_files_len = 0



def pause_console():
	os.system('pause')

def clear_console():
	os.system('cls')
	os.system('clear')

def print_program_announcement():
	global PROGRAM_ANNOUNCEMENT_STRING
	print(PROGRAM_ANNOUNCEMENT_STRING)

# 공무일지.xlsx 이 존재하는지 확인
def validate_target_file(target_file):
	current_path = os.getcwd()
	for f in os.listdir(current_path):
		if f.strip() == target_file:
			return True

	return False

def set_date_on_file(target_file):
	current_path = os.getcwd()
	target_file_name, target_file_extension = target_file.split('.')
	copyfile(target_file, "TEMP")

	date_string = date.today().strftime("%y-%m-%d")
	new_file_name = target_file_name + " " + date_string + "." + target_file_extension

	new_file_name_tmp = new_file_name
	new_file_index = 0
	duplicated = True
	while duplicated:
		again = False
		for f in os.listdir(current_path):
			if f.strip() == new_file_name_tmp:
				new_file_index += 1
				new_file_name_tmp = f"{new_file_name[:-5]} ({new_file_index}).xlsx"
				again = True
				break
		if again: continue

		duplicated = False

	new_file_name = new_file_name_tmp
	os.rename("TEMP", new_file_name)

	return new_file_name


# 공무일지 엑셀 템플릿
def set_date_in_file_workLogExcelFormat(target_file):
	# 기존 엑셀 파일 열기
	wb = load_workbook(target_file)

	# 활성화된 워크시트 가져오기
	ws = wb.active

	date_string = date.today().strftime("%Y년 %m월 %d일")
	weekday_string = get_korean_weekday()
	ws.cell(row=3, column=1, value=date_string + " " + weekday_string)
	
	wb.save(target_file)

def get_korean_weekday():
	weekdays = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]
	return weekdays[date.today().weekday()]


if __name__ == '__main__':
	while True:
		print_program_announcement()

		target_file = "공무일지.xlsx"
		if not validate_target_file(target_file):
			print(f"[시스템] {target_file} 파일이 존재하지 않습니다.")
			pause_console()
			break

		new_file = set_date_on_file(target_file)
		set_date_in_file_workLogExcelFormat(new_file)

		print(f"[시스템] {new_file} 파일로 변환 완료했습니다")
		pause_console()
		break
