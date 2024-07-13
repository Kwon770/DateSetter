import os
from os.path import isfile, join
from pathlib import Path
import sys
import re
from shutil import copyfile
from datetime import date
from openpyxl import load_workbook

PROGRAM_ANNOUNCEMENT_STRING = """
[ 날짜 입력기 ]

날짜를 입력할 파일 목록 입니다."""
EXCEL_EXTENSIONS = ["xlsx", "xlsm", "xlsb", "xltx", "xls", "xlt", "xml", "xlam", "xlw", "xlr"]

current_path = ""
excel_files_len = 0

def pause_console():
	os.system('pause')

def clear_console():
	os.system('cls')
	os.system('clear')

def print_program_announcement():
	global PROGRAM_ANNOUNCEMENT_STRING
	print(PROGRAM_ANNOUNCEMENT_STRING)


def scan_directory():
	global EXCEL_EXTENSIONS
	global current_path
	global excel_files_len

	current_path = Path.cwd()
	excel_file_list = []
	for f in os.listdir(current_path):
		if not isfile(join(current_path, f)):
			continue

		file_info = f.split('.')
		if len(file_info) != 2 or file_info[1] not in EXCEL_EXTENSIONS:
			continue

		file_info2 = re.split('[-.]', f)
		if len(file_info2) >= 3 and file_info2[-2].isnumeric() and file_info2[-3].isnumeric():
			continue

		excel_file_list.append(f)

	excel_files_len = len(excel_file_list)
	return excel_file_list



def print_files_list(files_list):
	print('- 0. 파일 목록 새로고침')
	for idx, file in enumerate(files_list):
		print(f'- {idx+1}. {file}')
	print()



def input_target_index():
	global excel_files_len

	while True:
		try:
			target_index = int(input("[입력] 변경할 파일의 번호를 입력하세요 : "))
			if target_index > excel_files_len:
				raise ValueError

			return target_index-1

		except ValueError:
			print("[에러] 올바르지 않은 입력 형태입니다.")


def validate_target_file(target_file):
	target_file_name, target_file_extension = target_file.split('.')
	date_string = date.today().strftime("%y-%m-%d")
	date_setter_name = target_file_name + " " + date_string + "." + target_file_extension
	for f in os.listdir(current_path):
		if f == date_setter_name:
			return False

	return True

def set_date_on_file(target_file):
	target_file_name, target_file_extension = target_file.split('.')
	copyfile(target_file, "TEMP")

	date_string = date.today().strftime("%y-%m-%d")
	new_file_name = target_file_name + " " + date_string + "." + target_file_extension
	os.rename("TEMP", new_file_name)

	return new_file_name


# 공무일지 엑셀 템플릿
def set_date_in_file_workLogExcelFormat(target_file):
	# 기존 엑셀 파일 열기
	wb = load_workbook(target_file)

	# 활성화된 워크시트 가져오기
	ws = wb.active

	date_string = date.today().strftime("%y년 %m월 %d일")
	weekday_string = get_korean_weekday()
	ws.cell(row=3, column=1, value=date_string + " " + weekday_string)
	
	wb.save(target_file)

def get_korean_weekday():
	weekdays = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]
	return weekdays[date.today().weekday()]


if __name__ == '__main__':
	while True:
		print_program_announcement()

		files_list = scan_directory()
		print_files_list(files_list)

		target_index = input_target_index()
		if target_index == -1:
			break

		target_file = files_list[target_index]
		if not validate_target_file(target_file):
			print("[시스템] 중복된 파일이 있습니다.")
			pause_console()
			continue

		new_file = set_date_on_file(target_file)
		set_date_in_file_workLogExcelFormat(new_file)
